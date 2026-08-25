import streamlit as st
import xml.etree.ElementTree as ET
from datetime import datetime
from pathlib import Path
import json
import openpyxl
from io import BytesIO
import pandas as pd

NS = {"nfe": "http://www.portalfiscal.inf.br/nfe"}

CONFIG_PATH = Path(__file__).parent / "cfop_config.json"
TEMPLATE_PATH = Path(__file__).parent / "MODELO_PARCELA.xlsx"

DEFAULT_CFOPS_EXCLUIDOS = [
    {"cfop": "1949", "descricao": "Outra entrada de mercadoria não especificada"},
    {"cfop": "2949", "descricao": "Outra entrada de mercadoria não especificada"},
    {"cfop": "1152", "descricao": "Transferência de mercadoria adquirida de terceiros"},
    {"cfop": "2152", "descricao": "Transferência de mercadoria adquirida de terceiros"},
    {"cfop": "1908", "descricao": "Remessa por conta e ordem de terceiros"},
    {"cfop": "1910", "descricao": "Entrada de bonificação, doação ou brinde"},
    {"cfop": "2910", "descricao": "Entrada de bonificação, doação ou brinde"},
]


# ---------------------------------------------------------------------------
# Configuração de CFOPs excluídos (persistida em disco)
# ---------------------------------------------------------------------------
def carregar_cfops_excluidos():
    if CONFIG_PATH.exists():
        try:
            return json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
        except Exception:
            pass
    return DEFAULT_CFOPS_EXCLUIDOS.copy()


def salvar_cfops_excluidos(lista):
    try:
        CONFIG_PATH.write_text(
            json.dumps(lista, ensure_ascii=False, indent=2), encoding="utf-8"
        )
    except Exception:
        # Em alguns ambientes hospedados o disco pode ser somente leitura ou
        # efêmero — nesse caso a configuração continua valendo para a sessão
        # atual (em st.session_state), só não sobrevive a um reinício do servidor.
        pass


# ---------------------------------------------------------------------------
# Parsing do XML da NF-e
# ---------------------------------------------------------------------------
def _find_text(elem, path):
    node = elem.find(path, NS)
    return node.text.strip() if node is not None and node.text else None


def parse_nfe(xml_bytes, filename):
    """Retorna um dicionário com os dados relevantes de uma NF-e, ou levanta ValueError."""
    root = ET.fromstring(xml_bytes)

    inf_nfe = root.find(".//nfe:infNFe", NS)
    if inf_nfe is None:
        raise ValueError("Não encontrei a tag infNFe — arquivo não parece ser uma NF-e válida.")

    ide = inf_nfe.find("nfe:ide", NS)
    emit = inf_nfe.find("nfe:emit", NS)
    total = inf_nfe.find(".//nfe:ICMSTot", NS)
    cobr = inf_nfe.find("nfe:cobr", NS)

    n_nf = _find_text(ide, "nfe:nNF")
    dh_emi = _find_text(ide, "nfe:dhEmi") or _find_text(ide, "nfe:dEmi")
    data_emissao = None
    if dh_emi:
        data_emissao = dh_emi[:10]  # AAAA-MM-DD, funciona tanto para dhEmi quanto dEmi

    nome_fornecedor = _find_text(emit, "nfe:xNome")
    v_nf = _find_text(total, "nfe:vNF")

    # CFOPs presentes na nota (um por item)
    cfops = set()
    for det in inf_nfe.findall("nfe:det", NS):
        cfop = _find_text(det, "nfe:prod/nfe:CFOP")
        if cfop:
            cfops.add(cfop)

    # Parcelas (duplicatas)
    parcelas = []
    if cobr is not None:
        for dup in cobr.findall("nfe:dup", NS):
            v_dup = _find_text(dup, "nfe:vDup")
            d_venc = _find_text(dup, "nfe:dVenc")
            if v_dup:
                parcelas.append({"vencimento": d_venc, "valor": float(v_dup)})

    return {
        "arquivo": filename,
        "numero_nf": n_nf,
        "fornecedor": nome_fornecedor,
        "data_emissao": data_emissao,
        "valor_total": float(v_nf) if v_nf else None,
        "cfops": cfops,
        "parcelas": parcelas,
    }


def formatar_data_br(data_iso):
    if not data_iso:
        return None
    try:
        return datetime.strptime(data_iso, "%Y-%m-%d").strftime("%d/%m/%Y")
    except ValueError:
        return data_iso


def processar_notas(arquivos, cfops_excluidos):
    """Processa a lista de arquivos XML e devolve (linhas, ignoradas, avisos, erros)."""
    excluidos = {c["cfop"] for c in cfops_excluidos}

    linhas = []       # linhas prontas para a planilha final
    ignoradas = []     # notas 100% com CFOP excluído -> não geram parcela
    avisos = []        # notas com CFOP misto (alguns excluídos, outros não)
    erros = []         # arquivos que não puderam ser lidos

    for arq in arquivos:
        try:
            dados = parse_nfe(arq.getvalue(), arq.name)
        except Exception as e:
            erros.append({"arquivo": arq.name, "erro": str(e)})
            continue

        cfops = dados["cfops"]
        if cfops and cfops.issubset(excluidos):
            ignoradas.append({
                "arquivo": dados["arquivo"],
                "numero_nf": dados["numero_nf"],
                "fornecedor": dados["fornecedor"],
                "cfops": ", ".join(sorted(cfops)),
            })
            continue

        if cfops and (cfops & excluidos) and not cfops.issubset(excluidos):
            avisos.append({
                "arquivo": dados["arquivo"],
                "numero_nf": dados["numero_nf"],
                "fornecedor": dados["fornecedor"],
                "cfops": ", ".join(sorted(cfops)),
                "motivo": "Nota tem CFOPs excluídos e não excluídos misturados — foi processada por inteiro, confira.",
            })

        parcelas = dados["parcelas"]
        if not parcelas:
            # Fallback: sem parcela no XML -> 1 parcela, vencimento = data de emissão
            parcelas = [{
                "vencimento": dados["data_emissao"],
                "valor": dados["valor_total"],
            }]

        for p in parcelas:
            linhas.append({
                "NUMERO NF": dados["numero_nf"],
                "NOME DO FORNECEDOR": dados["fornecedor"],
                "VALOR PARCELA": p["valor"],
                "VENCIMENTO": formatar_data_br(p["vencimento"]),
            })

    return linhas, ignoradas, avisos, erros


def gerar_planilha(linhas):
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb.active
    row_idx = 2
    for linha in linhas:
        ws.cell(row=row_idx, column=1, value=linha["NUMERO NF"])
        ws.cell(row=row_idx, column=2, value=linha["NOME DO FORNECEDOR"])
        ws.cell(row=row_idx, column=3, value=linha["VALOR PARCELA"])
        ws.cell(row=row_idx, column=4, value=linha["VENCIMENTO"])
        row_idx += 1
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ---------------------------------------------------------------------------
# Proteção por senha
# ---------------------------------------------------------------------------
st.set_page_config(page_title="Parcelas NF-e → Domínio", layout="wide")


def checar_senha():
    """Mostra uma tela de login simples. Retorna True se a senha estiver correta."""
    senha_configurada = st.secrets.get("app_password")

    if not senha_configurada:
        # Nenhuma senha configurada (ex.: primeira instalação) — avisa e libera,
        # para não travar o uso local antes de configurar o secrets.toml.
        st.warning(
            "⚠️ Nenhuma senha configurada (app_password ausente em secrets.toml). "
            "O app está liberado sem proteção. Veja o README para configurar a senha."
        )
        return True

    if st.session_state.get("autenticado"):
        return True

    st.title("🔒 Parcelas NF-e → Domínio")
    senha_digitada = st.text_input("Senha de acesso", type="password")
    if st.button("Entrar"):
        if senha_digitada == senha_configurada:
            st.session_state.autenticado = True
            st.rerun()
        else:
            st.error("Senha incorreta.")
    return False


if not checar_senha():
    st.stop()


# ---------------------------------------------------------------------------
# Interface
# ---------------------------------------------------------------------------
st.title("📑 Extração de Parcelas de NF-e para importação no Domínio")
st.caption(
    "Envie os XMLs das notas de entrada. O app extrai as parcelas (duplicatas) "
    "e gera a planilha no layout exigido pelo Domínio."
)

with st.sidebar:
    st.header("⚙️ Configurações")
    st.subheader("CFOPs que NÃO geram parcela")
    st.caption("Ex.: outras entradas, transferências, bonificação, remessas etc.")

    if "cfops_excluidos" not in st.session_state:
        st.session_state.cfops_excluidos = carregar_cfops_excluidos()

    df_cfops = pd.DataFrame(st.session_state.cfops_excluidos)
    df_editado = st.data_editor(
        df_cfops,
        num_rows="dynamic",
        use_container_width=True,
        column_config={
            "cfop": st.column_config.TextColumn("CFOP", required=True),
            "descricao": st.column_config.TextColumn("Descrição"),
        },
        key="editor_cfops",
    )

    if st.button("💾 Salvar configuração de CFOPs"):
        lista = df_editado.dropna(subset=["cfop"]).to_dict("records")
        st.session_state.cfops_excluidos = lista
        salvar_cfops_excluidos(lista)
        st.success("Configuração salva para esta sessão.")

    st.divider()
    st.caption(
        "⚠️ No app hospedado online (Streamlit Cloud), o servidor pode reiniciar "
        "e apagar alterações salvas apenas em disco. Para não perder sua lista de "
        "CFOPs, baixe o arquivo abaixo depois de editar, e carregue-o novamente "
        "sempre que abrir o app (ou substitua o cfop_config.json no repositório do "
        "GitHub para tornar a mudança permanente)."
    )

    lista_atual = df_editado.dropna(subset=["cfop"]).to_dict("records")
    st.download_button(
        "⬇️ Baixar configuração atual (JSON)",
        data=json.dumps(lista_atual, ensure_ascii=False, indent=2),
        file_name="cfop_config.json",
        mime="application/json",
    )

    config_upload = st.file_uploader(
        "⬆️ Carregar configuração salva (JSON)", type=["json"], key="upload_cfop_config"
    )
    if config_upload is not None:
        try:
            nova_lista = json.loads(config_upload.getvalue().decode("utf-8"))
            st.session_state.cfops_excluidos = nova_lista
            salvar_cfops_excluidos(nova_lista)
            st.success("Configuração carregada. Role para cima para conferir a tabela.")
            st.rerun()
        except Exception as e:
            st.error(f"Não consegui ler esse arquivo: {e}")

st.subheader("1. Envie os XMLs das notas de entrada")
arquivos = st.file_uploader(
    "Arquivos XML", type=["xml"], accept_multiple_files=True
)

if arquivos:
    st.subheader("2. Resultado")
    linhas, ignoradas, avisos, erros = processar_notas(
        arquivos, st.session_state.cfops_excluidos
    )

    col1, col2, col3, col4 = st.columns(4)
    col1.metric("Notas enviadas", len(arquivos))
    col2.metric("Parcelas geradas", len(linhas))
    col3.metric("Notas ignoradas (CFOP)", len(ignoradas))
    col4.metric("Erros de leitura", len(erros))

    if linhas:
        st.markdown("**Prévia das parcelas:**")
        st.dataframe(pd.DataFrame(linhas), use_container_width=True)

        planilha = gerar_planilha(linhas)
        st.download_button(
            "⬇️ Baixar planilha para importar no Domínio",
            data=planilha,
            file_name="parcelas_importacao_dominio.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    else:
        st.warning("Nenhuma parcela foi gerada com os arquivos enviados.")

    if avisos:
        st.markdown("**⚠️ Notas com CFOP misto (revisar):**")
        st.dataframe(pd.DataFrame(avisos), use_container_width=True)

    if ignoradas:
        with st.expander(f"🚫 Notas ignoradas por CFOP ({len(ignoradas)})"):
            st.dataframe(pd.DataFrame(ignoradas), use_container_width=True)

    if erros:
        with st.expander(f"❌ Arquivos com erro de leitura ({len(erros)})"):
            st.dataframe(pd.DataFrame(erros), use_container_width=True)
else:
    st.info("Envie um ou mais arquivos XML para começar.")
